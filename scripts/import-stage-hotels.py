import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT = ROOT / "assets" / "data" / "hotels.json"
SHEET_NAME = "全国料金付き一覧"


def clean(value):
    if value is None:
        return ""
    if hasattr(value, "isoformat"):
        value = value.isoformat()
    return unicodedata.normalize("NFKC", str(value)).replace("。", "").strip()


def number(value):
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def extract_identifier(url, patterns):
    text = clean(url)
    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            return match.group(1)
    return None


def compact(values):
    return [value for value in values if value]


def build_hotel(row):
    venue_number = number(row[0])
    rank = number(row[4])
    if not venue_number or not rank:
        return None

    prefecture = clean(row[1])
    venue_name = clean(row[2])
    venue_area = clean(row[3])
    hotel_name = clean(row[5])
    nearest = clean(row[6])
    access = clean(row[7])
    adopted_price = clean(row[10])
    price_range = clean(row[11])
    plan = clean(row[12])
    room = clean(row[13])
    cancellation = clean(row[14])
    researched_at = clean(row[15])
    accuracy = clean(row[16])
    reason = clean(row[17])
    rakuten_source = clean(row[18])
    jalan_source = clean(row[19])

    if not venue_name or not hotel_name:
        return None

    venue_id = f"stage-{venue_number:03d}"
    hotel_id = f"{venue_id}-{rank}"
    priorities = ["station"]
    if re.search(r"徒歩|直結|すぐ|隣", access):
        priorities.insert(0, "near")

    facts = compact([
        access,
        f"料金目安 {adopted_price}" if adopted_price else "",
        f"料金幅 {price_range}" if price_range else "",
    ])[:3]

    statuses = []
    if access:
        statuses.append(["会場へのアクセス", "check", access])
    if adopted_price:
        statuses.append(["料金目安", "check", adopted_price])

    hotel = {
        "id": hotel_id,
        "genre": "stage",
        "name": hotel_name,
        "role": f"{venue_name}周辺の候補{rank}",
        "area": venue_area or prefecture,
        "prefecture": prefecture,
        "venueLabel": venue_name,
        "venueId": venue_id,
        "venueNumber": venue_number,
        "rank": rank,
        "venues": [venue_id],
        "station": nearest or "最寄り情報は予約前に確認",
        "accessEstimate": access,
        "summary": reason or f"{venue_name}への移動候補として比較できます",
        "styles": ["solo", "group", "matinee", "stay"],
        "priorities": priorities,
        "facts": facts,
        "statuses": statuses,
        "priceEstimate": adopted_price,
        "priceRangeEstimate": price_range,
        "rakutenPriceEstimate": clean(row[8]),
        "jalanPriceEstimate": clean(row[9]),
        "priceNumeric": number(row[24]),
        "planEstimate": plan,
        "roomTypeEstimate": room,
        "cancellationEstimate": cancellation,
        "researchedAt": researched_at,
        "accuracy": accuracy,
        "calculationReason": reason,
        "hotelMapUrl": clean(row[20]),
        "venueMapUrl": clean(row[21]),
        "venueOfficialUrl": clean(row[22]),
        "venueMemo": clean(row[23]),
        "rakutenHotelNo": extract_identifier(
            rakuten_source,
            [r"/HOTEL/(\d+)", r"/hotelinfo/(?:plan/)?(\d+)", r"[?&](?:hotelNo|f_no)=(\d+)"],
        ),
        "jalanYadNo": extract_identifier(jalan_source, [r"/yad(\d+)", r"[?&](?:yadNo|yad_no)=(\d+)"]),
        "isPublished": True,
    }
    return hotel


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: import-stage-hotels.py <xlsx-path>")

    workbook_path = Path(sys.argv[1])
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    existing = json.loads(OUTPUT.read_text(encoding="utf-8"))
    preserved = [hotel for hotel in existing.get("hotels", []) if hotel.get("genre") != "stage"]

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise SystemExit(f"Sheet not found: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    imported = []
    for row in sheet.iter_rows(min_row=2, max_col=25, values_only=True):
        hotel = build_hotel(row)
        if hotel:
            imported.append(hotel)

    venue_ids = {hotel["venueId"] for hotel in imported}
    hotel_ids = {hotel["id"] for hotel in imported}
    if len(imported) != 1250 or len(venue_ids) != 250 or len(hotel_ids) != 1250:
        raise SystemExit(
            f"Unexpected counts: hotels={len(imported)}, venues={len(venue_ids)}, ids={len(hotel_ids)}"
        )

    payload = {
        "schemaVersion": 2,
        "generatedFrom": workbook_path.name,
        "stageVenueCount": len(venue_ids),
        "stageHotelCount": len(imported),
        "hotels": imported + preserved,
    }
    OUTPUT.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n", encoding="utf-8")

    rakuten_ids = sum(bool(hotel.get("rakutenHotelNo")) for hotel in imported)
    jalan_ids = sum(bool(hotel.get("jalanYadNo")) for hotel in imported)
    print(
        json.dumps(
            {
                "stageHotels": len(imported),
                "stageVenues": len(venue_ids),
                "preservedHotels": len(preserved),
                "rakutenIdentifiers": rakuten_ids,
                "jalanIdentifiers": jalan_ids,
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()
